"""Consolidate audit JSONs from the 4 sub-agents into a single review spreadsheet.

Reads:
    planning/audits/audit-1-boundary-a.json
    planning/audits/audit-2-boundary-b.json
    planning/audits/audit-3a-hardgate-scope.json
    planning/audits/audit-3b-false-negatives.json

Writes:
    planning/audits/consolidated.xlsx with 4 sheets:
        - flips_high_confidence : confidence == high in ≥1 agent, no contradiction
        - flips_review          : medium/low confidence, agents agree on direction
        - conflicts             : ≥1 agent flips, ≥1 other agent kept (borderline)
        - borderline_kept       : visibility only — all agents kept

Usage:
    python scripts/consolidate_audits.py
"""

import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
AUDITS_DIR = ROOT / "planning" / "audits"
XLSX = ROOT / "Análise.xlsx"
OUTPUT = AUDITS_DIR / "consolidated.xlsx"

AUDIT_FILES = {
    "1-boundary-a":  AUDITS_DIR / "audit-1-boundary-a.json",
    "2-boundary-b":  AUDITS_DIR / "audit-2-boundary-b.json",
    "3a-hardgate":   AUDITS_DIR / "audit-3a-hardgate-scope.json",
    "3b-false-neg":  AUDITS_DIR / "audit-3b-false-negatives.json",
}

CONF_RANK = {"high": 3, "medium": 2, "low": 1}

GREEN = PatternFill(start_color="E7FFEE", end_color="E7FFEE", fill_type="solid")
RED = PatternFill(start_color="FFEFF0", end_color="FFEFF0", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)

HEADERS = [
    "row", "paper_id", "title", "year",
    "current_decision", "recommended_decision",
    "source_agents", "confidence_max", "num_flips", "num_borderline",
    "union_rationale", "evidence_quotes",
    "Approved by Reviewer",
]


def plain(v):
    """Extract plain text from a cell value (str or CellRichText)."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    try:
        return "".join(p.text if hasattr(p, "text") else p for p in v)
    except Exception:
        return str(v)


def load_audits():
    audits = {}
    for key, path in AUDIT_FILES.items():
        if path.exists():
            with open(path, encoding="utf-8") as f:
                audits[key] = json.load(f)
        else:
            print(f"WARN: missing {path}")
            audits[key] = {"flips": [], "borderline_kept": []}
    return audits


def load_paper_meta():
    """Return dict paper_id -> {row, title, year, current_decision}."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["05 — Abstract Screening"]
    meta = {}
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, 1).value
        if not pid:
            continue
        title = plain(ws.cell(r, 2).value)
        year = ws.cell(r, 3).value
        auto = ws.cell(r, 5).value
        myfin = ws.cell(r, 9).value
        effective = myfin if myfin in ("Include", "Exclude") else auto
        meta[pid] = dict(row=r, title=title, year=year, current=effective)
    return meta


def aggregate(audits):
    """Group items by paper_id. Returns dict pid -> {flips, borderline}."""
    by_paper = defaultdict(lambda: {"flips": [], "borderline": []})
    for agent, audit in audits.items():
        for item in audit.get("flips", []) or []:
            entry = dict(item)
            entry["agent"] = agent
            by_paper[entry["paper_id"]]["flips"].append(entry)
        for item in audit.get("borderline_kept", []) or []:
            entry = dict(item)
            entry["agent"] = agent
            by_paper[entry["paper_id"]]["borderline"].append(entry)
    return by_paper


def build_row(pid, meta, flips, bord):
    flips_sorted = sorted(flips, key=lambda f: -CONF_RANK.get(f.get("confidence", "low"), 0))
    rec = flips_sorted[0]["recommended_decision"] if flips_sorted else meta["current"]
    union_lines = [f"[{f['agent']}] {f.get('rationale_pt', '')}" for f in flips]
    union_lines += [f"[{b['agent']}/keep] {b.get('rationale_pt', '')}" for b in bord]
    quotes = [f.get("evidence_quote", "") for f in flips if f.get("evidence_quote")]
    return {
        "row": meta["row"],
        "paper_id": pid,
        "title": meta["title"],
        "year": meta["year"],
        "current_decision": meta["current"],
        "recommended_decision": rec,
        "source_agents": ", ".join(sorted({f["agent"] for f in flips + bord})),
        "confidence_max": flips_sorted[0].get("confidence", "—") if flips_sorted else "—",
        "num_flips": len(flips),
        "num_borderline": len(bord),
        "union_rationale": " | ".join(union_lines),
        "evidence_quotes": " | ".join(quotes),
    }


def categorize(by_paper, meta):
    high_conf, review, conflict, borderline_only = [], [], [], []
    for pid, items in by_paper.items():
        flips = items["flips"]
        bord = items["borderline"]
        m = meta.get(pid, dict(row=None, title="(unknown)", year=None, current="?"))

        if not flips:
            if bord:
                row = build_row(pid, m, [], bord)
                borderline_only.append(row)
            continue

        # ≥1 agent recommended a flip.
        if bord:
            # Another agent kept it as borderline_kept — true disagreement.
            conflict.append(build_row(pid, m, flips, bord))
            continue

        max_conf = max(CONF_RANK.get(f.get("confidence", "low"), 1) for f in flips)
        row = build_row(pid, m, flips, bord)
        if max_conf == CONF_RANK["high"]:
            high_conf.append(row)
        else:
            review.append(row)

    # Sort each category by row
    sort_key = lambda x: x["row"] if x["row"] is not None else 9999
    return (
        sorted(high_conf, key=sort_key),
        sorted(review, key=sort_key),
        sorted(conflict, key=sort_key),
        sorted(borderline_only, key=sort_key),
    )


def write_sheet(ws, title, rows, default_approved):
    ws.title = title
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(HEADERS[:-1], start=1):
            ws.cell(r, c, row.get(key, ""))
        ws.cell(r, len(HEADERS), default_approved if default_approved else "")
        rec = row.get("recommended_decision", "")
        fill = GREEN if rec == "Include" else (RED if rec == "Exclude" else None)
        if fill:
            ws.cell(r, 2).fill = fill  # paper_id
            ws.cell(r, 3).fill = fill  # title
    widths = {1: 6, 2: 14, 3: 60, 4: 6, 5: 10, 6: 12, 7: 24, 8: 10, 9: 6, 10: 6, 11: 80, 12: 60, 13: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "C2"


def main():
    AUDITS_DIR.mkdir(parents=True, exist_ok=True)
    audits = load_audits()
    meta = load_paper_meta()
    by_paper = aggregate(audits)
    high_conf, review, conflict, borderline_only = categorize(by_paper, meta)

    print(f"flips_high_confidence: {len(high_conf)}  (default Approved=TRUE)")
    print(f"flips_review:          {len(review)}  (default empty — review needed)")
    print(f"conflicts:             {len(conflict)}  (default empty — review needed)")
    print(f"borderline_kept:       {len(borderline_only)}  (visibility only)")

    wb = openpyxl.Workbook()
    write_sheet(wb.active, "flips_high_confidence", high_conf, default_approved="TRUE")
    write_sheet(wb.create_sheet(), "flips_review", review, default_approved="")
    write_sheet(wb.create_sheet(), "conflicts", conflict, default_approved="")
    write_sheet(wb.create_sheet(), "borderline_kept", borderline_only, default_approved="")
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")


if __name__ == "__main__":
    main()
