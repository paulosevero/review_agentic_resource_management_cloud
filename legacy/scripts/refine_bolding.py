"""Refines title/abstract bolding in Análise.xlsx — strips editorial noise
and re-applies bold to a curated list of keywords using spaCy PhraseMatcher.

Pipeline:
1. Load curated keywords + strip patterns from protocols/bolding-keywords.yaml.
2. For each row in the target sheet:
   a. Strip editorial noise from Title and Abstract (regex substitutions).
   b. Run spaCy PhraseMatcher (LOWER attr) over the cleaned text.
   c. Merge overlapping/adjacent spans (separator = whitespace or hyphen).
   d. Trim leading/trailing stopwords from each span.
   e. Apply density cap (≤ 50% of characters bolded).
   f. Build openpyxl CellRichText with bold runs.
3. Dry-run mode prints diffs for the first N rows.
4. Apply mode overwrites cells in place.

Usage:
    python scripts/refine_bolding.py --dry-run --sample 8
    python scripts/refine_bolding.py --apply
"""

import argparse
import re
import sys
from pathlib import Path

import yaml
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

import spacy
from spacy.matcher import PhraseMatcher

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Análise.xlsx"
KEYWORDS_FILE = ROOT / "protocols" / "bolding-keywords.yaml"

SHEET = "05 — Abstract Screening"
COL_TITLE = 2  # B
COL_ABSTRACT = 4  # D

DENSITY_CAP = 0.50  # max fraction of characters in bold

# Hyphen-or-whitespace gap for merging adjacent bold spans
GAP_RE = re.compile(r"^[\s\-]+$")


def load_config():
    with open(KEYWORDS_FILE, encoding="utf-8") as f:
        return yaml.safe_load(f)


def compile_strip_patterns(strip_cfg):
    compiled = []
    flag_map = {"DOTALL": re.DOTALL, "MULTILINE": re.MULTILINE, "IGNORECASE": re.IGNORECASE}
    for entry in strip_cfg:
        flags = re.IGNORECASE  # default case-insensitive
        for fname in entry.get("flags", []):
            flags |= flag_map[fname]
        compiled.append({
            "regex": re.compile(entry["pattern"], flags),
            "replace": entry.get("replace", ""),
        })
    return compiled


def strip_noise(text, compiled_patterns):
    if not text:
        return ""
    out = text
    for entry in compiled_patterns:
        out = entry["regex"].sub(entry["replace"], out)
    return out.strip()


def build_matcher(nlp, keywords_groups):
    matcher = PhraseMatcher(nlp.vocab, attr="LOWER")
    seen = set()
    for group, terms in keywords_groups.items():
        patterns = []
        for term in terms:
            t = term.strip().lower()
            if not t or t in seen:
                continue
            seen.add(t)
            patterns.append(nlp.make_doc(term))
        if patterns:
            matcher.add(group, patterns)
    return matcher


def find_spans(text, nlp, matcher, stopwords):
    """Return list of (start, end) char offsets in `text` that should be bold."""
    if not text.strip():
        return []
    doc = nlp(text)
    raw = []
    for _match_id, start, end in matcher(doc):
        span = doc[start:end]
        raw.append((span.start_char, span.end_char))
    # 1. Sort and merge overlapping
    raw.sort()
    merged = []
    for s, e in raw:
        if merged and s <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(merged[-1][1], e))
        else:
            merged.append((s, e))
    # 2. Merge adjacent spans separated only by hyphen/whitespace (e.g., multi-LLMs)
    fused = []
    for s, e in merged:
        if fused:
            gap = text[fused[-1][1]:s]
            if gap and GAP_RE.match(gap):
                fused[-1] = (fused[-1][0], e)
                continue
        fused.append((s, e))
    # 3. Extend right for plural "s" if directly attached and not already covered
    extended = []
    for s, e in fused:
        if e < len(text) and text[e:e+1] == "s" and (e+1 == len(text) or not text[e+1].isalpha()):
            e += 1
        extended.append((s, e))
    # 4. Trim leading/trailing stopwords and punctuation
    trimmed = []
    for s, e in extended:
        # trim left
        while s < e:
            # skip punctuation/whitespace
            if not text[s].isalnum():
                s += 1
                continue
            # check leading stopword
            tail = text[s:e]
            m = re.match(r"([A-Za-z][A-Za-z\-']*)", tail)
            if not m:
                break
            tok = m.group(1).lower()
            if tok in stopwords:
                s += len(m.group(1))
                # skip following whitespace
                while s < e and not text[s].isalnum():
                    s += 1
                continue
            break
        # trim right
        while s < e:
            if not text[e-1].isalnum():
                e -= 1
                continue
            head = text[s:e]
            m = re.search(r"([A-Za-z][A-Za-z\-']*)$", head)
            if not m:
                break
            tok = m.group(1).lower()
            if tok in stopwords:
                e -= len(m.group(1))
                while s < e and not text[e-1].isalnum():
                    e -= 1
                continue
            break
        if e > s:
            trimmed.append((s, e))
    # 5. Density cap: drop the shortest spans until below cap
    total = sum(e - s for s, e in trimmed)
    if total > DENSITY_CAP * max(len(text), 1):
        trimmed.sort(key=lambda x: x[1] - x[0], reverse=True)
        kept = []
        running = 0
        for s, e in trimmed:
            if running + (e - s) > DENSITY_CAP * len(text):
                continue
            kept.append((s, e))
            running += (e - s)
        trimmed = kept
        trimmed.sort()
    return trimmed


def build_rich_text(text, spans):
    """Return CellRichText with bold TextBlocks at the given spans."""
    if not spans:
        return text
    bold_font = InlineFont(b=True)
    parts = []
    cursor = 0
    for s, e in spans:
        if s > cursor:
            parts.append(text[cursor:s])
        parts.append(TextBlock(font=bold_font, text=text[s:e]))
        cursor = e
    if cursor < len(text):
        parts.append(text[cursor:])
    return CellRichText(parts)


def render_span_preview(text, spans):
    """Return text with **markers** around bold spans (for terminal preview)."""
    if not spans:
        return text
    out = []
    cursor = 0
    for s, e in spans:
        out.append(text[cursor:s])
        out.append("**" + text[s:e] + "**")
        cursor = e
    out.append(text[cursor:])
    return "".join(out)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Preview changes for first N rows")
    parser.add_argument("--apply", action="store_true", help="Apply changes to spreadsheet")
    parser.add_argument("--sample", type=int, default=8, help="Number of rows to preview in dry-run")
    parser.add_argument("--rows", type=str, default=None,
                        help="Comma-separated row numbers to preview (overrides --sample)")
    args = parser.parse_args()

    if not (args.dry_run or args.apply):
        parser.error("Pass --dry-run or --apply")

    cfg = load_config()
    keywords_groups = cfg["keywords"]
    strip_patterns = compile_strip_patterns(cfg["strip_patterns"])
    stopwords = set(str(w).lower() for w in cfg["stopword_trim"])

    print(f"Loading spaCy model en_core_web_sm ...", file=sys.stderr)
    try:
        nlp = spacy.load("en_core_web_sm", disable=["ner", "parser", "tagger", "lemmatizer"])
    except OSError:
        print("ERROR: en_core_web_sm not installed. Run: python -m spacy download en_core_web_sm", file=sys.stderr)
        sys.exit(1)
    matcher = build_matcher(nlp, keywords_groups)
    n_terms = sum(len(v) for v in keywords_groups.values())
    print(f"Loaded {n_terms} keyword patterns across {len(keywords_groups)} groups", file=sys.stderr)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    if args.rows:
        target_rows = [int(x) for x in args.rows.split(",")]
    else:
        target_rows = list(range(2, ws.max_row + 1))

    preview_count = 0
    apply_count = 0

    for r in target_rows:
        title = ws.cell(r, COL_TITLE).value
        abstract = ws.cell(r, COL_ABSTRACT).value
        # Cell value may already be CellRichText — extract plain text
        if hasattr(title, "__iter__") and not isinstance(title, str):
            title = "".join(p.text if hasattr(p, "text") else p for p in title)
        if hasattr(abstract, "__iter__") and not isinstance(abstract, str):
            abstract = "".join(p.text if hasattr(p, "text") else p for p in abstract)

        title = title or ""
        abstract = abstract or ""

        clean_title = strip_noise(title, strip_patterns)
        clean_abstract = strip_noise(abstract, strip_patterns)

        title_spans = find_spans(clean_title, nlp, matcher, stopwords)
        abstract_spans = find_spans(clean_abstract, nlp, matcher, stopwords)

        if args.dry_run:
            if preview_count >= args.sample and not args.rows:
                continue
            print(f"\n--- row {r} ---")
            print(f"TITLE    : {render_span_preview(clean_title, title_spans)}")
            print(f"ABSTRACT : {render_span_preview(clean_abstract[:500], [s for s in abstract_spans if s[1] <= 500])}")
            preview_count += 1
        else:
            ws.cell(r, COL_TITLE).value = build_rich_text(clean_title, title_spans)
            ws.cell(r, COL_ABSTRACT).value = build_rich_text(clean_abstract, abstract_spans)
            apply_count += 1

    if args.apply:
        wb.save(XLSX)
        print(f"\nApplied refined bolding to {apply_count} rows. Saved: {XLSX}", file=sys.stderr)
    else:
        print(f"\nDry-run complete ({preview_count} rows previewed). Use --apply to write.", file=sys.stderr)


if __name__ == "__main__":
    main()
