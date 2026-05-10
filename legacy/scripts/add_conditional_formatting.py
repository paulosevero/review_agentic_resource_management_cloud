"""Adds conditional formatting to a screening spreadsheet so the entire row
takes the green/red fill driven by `My Final Decision` (column H).

Bold-on-evidence in Title and Abstract cells is already applied at row-write
time by `apply_stage_05_to_spreadsheet.py` (CellRichText with TextBlock).
This script is idempotent — re-runs replace existing rules.
"""

import argparse
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

GREEN_FILL = PatternFill(start_color="FFE7FFEE", end_color="FFE7FFEE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFFFEFF0", end_color="FFFFEFF0", fill_type="solid")


def main(xlsx_path: str, sheet_name: str, decision_col_letter: str = "H") -> None:
    wb = openpyxl.load_workbook(xlsx_path, rich_text=True)
    ws = wb[sheet_name]
    last_col = ws.max_column
    last_row = ws.max_row
    full_range = f"A2:{get_column_letter(last_col)}{last_row}"
    ws.conditional_formatting._cf_rules.clear()
    include_rule = FormulaRule(formula=[f'${decision_col_letter}2="Include"'], fill=GREEN_FILL, stopIfTrue=True)
    exclude_rule = FormulaRule(formula=[f'${decision_col_letter}2="Exclude"'], fill=RED_FILL, stopIfTrue=True)
    ws.conditional_formatting.add(full_range, include_rule)
    ws.conditional_formatting.add(full_range, exclude_rule)
    wb.save(xlsx_path)
    print(f"Applied conditional formatting to {full_range} in '{sheet_name}'")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="spreadsheet_abstract_screening.xlsx")
    parser.add_argument("--sheet", default="05 — Abstract Screening")
    parser.add_argument("--decision-col", default="H")
    args = parser.parse_args()
    main(xlsx_path=args.xlsx, sheet_name=args.sheet, decision_col_letter=args.decision_col)
