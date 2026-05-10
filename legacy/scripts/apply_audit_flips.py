"""Apply approved flips from consolidated.xlsx to Análise.xlsx (Stage 05).

Reads planning/audits/consolidated.xlsx, filters rows in tabs
`flips_high_confidence`, `flips_review`, `conflicts` where
`Approved by Reviewer` is truthy, then updates `My Final Decision`,
`My Justification`, and ID/Title fill colors on `05 — Abstract Screening`.

Backups Análise.xlsx → Análise.audit-backup.xlsx before writing.
Logs every applied flip to planning/audits/applied-flips.md.

Usage:
    python scripts/apply_audit_flips.py --dry-run
    python scripts/apply_audit_flips.py --apply
"""

import argparse
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Análise.xlsx"
BACKUP = ROOT / "Análise.audit-backup.xlsx"
CONSOLIDATED = ROOT / "planning" / "audits" / "consolidated.xlsx"
LOG_FILE = ROOT / "planning" / "audits" / "applied-flips.md"

GREEN = PatternFill(start_color="E7FFEE", end_color="E7FFEE", fill_type="solid")
RED = PatternFill(start_color="FFEFF0", end_color="FFEFF0", fill_type="solid")

SHEET = "05 — Abstract Screening"
COL_ID = 1
COL_TITLE = 2
COL_MYDEC = 9
COL_MYJUST = 10

REVIEW_SHEETS = ["flips_high_confidence", "flips_review", "conflicts"]
TRUTHY = {"true", "yes", "y", "x", "1", "approved", "ok"}


def is_truthy(val):
    if val is None:
        return False
    return str(val).strip().lower() in TRUTHY


def collect_approved():
    if not CONSOLIDATED.exists():
        raise FileNotFoundError(f"Run consolidate_audits.py first — missing {CONSOLIDATED}")
    wb = openpyxl.load_workbook(CONSOLIDATED, data_only=True)
    approved = []
    for sheet_name in REVIEW_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1):
            row = {h: ws.cell(r, c).value for h, c in headers.items()}
            if not is_truthy(row.get("Approved by Reviewer")):
                continue
            row["_source_sheet"] = sheet_name
            approved.append(row)
    return approved


def apply_flips(approved):
    shutil.copy(XLSX, BACKUP)
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    pid_to_row = {ws.cell(r, COL_ID).value: r for r in range(2, ws.max_row + 1)}

    log_entries = []
    skipped = []
    for a in approved:
        pid = a.get("paper_id")
        r = pid_to_row.get(pid)
        if not r:
            skipped.append(pid)
            continue
        new_dec = a.get("recommended_decision")
        if new_dec not in ("Include", "Exclude"):
            skipped.append(f"{pid} (bad decision: {new_dec!r})")
            continue
        old_dec = ws.cell(r, COL_MYDEC).value or "(auto)"
        rationale = (a.get("union_rationale") or "").strip()
        agents = (a.get("source_agents") or "").strip()
        prefix = f"[Audit: {agents}]" if agents else "[Audit]"
        full_just = f"{prefix} {rationale}".strip()

        ws.cell(r, COL_MYDEC).value = new_dec
        ws.cell(r, COL_MYJUST).value = full_just
        fill = GREEN if new_dec == "Include" else RED
        ws.cell(r, COL_ID).fill = fill
        ws.cell(r, COL_TITLE).fill = fill

        log_entries.append(dict(
            row=r, pid=pid, old=str(old_dec), new=new_dec,
            agents=agents, rationale=rationale, source=a.get("_source_sheet"),
        ))

    wb.save(XLSX)
    return log_entries, skipped


def append_log(log_entries):
    if not log_entries:
        return
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    new_log = not LOG_FILE.exists()
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        if new_log:
            f.write("# Audit Flips Applied\n\n")
            f.write("Log of flips applied via `apply_audit_flips.py`. "
                    "Each section corresponds to one run.\n\n")
        f.write(f"## Run at {ts}\n\n")
        f.write(f"Applied {len(log_entries)} flips.\n\n")
        f.write("| Row | Paper ID | Old | New | Source Sheet | Agents | Rationale |\n")
        f.write("|---|---|---|---|---|---|---|\n")
        for e in log_entries:
            rat = (e["rationale"] or "").replace("|", "\\|").replace("\n", " ")
            if len(rat) > 240:
                rat = rat[:237] + "..."
            f.write(f"| {e['row']} | {e['pid']} | {e['old']} | {e['new']} | "
                    f"{e['source']} | {e['agents']} | {rat} |\n")
        f.write("\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Preview approved flips, no writes")
    parser.add_argument("--apply", action="store_true", help="Apply flips and write spreadsheet")
    args = parser.parse_args()
    if not (args.dry_run or args.apply):
        parser.error("Pass --dry-run or --apply")

    approved = collect_approved()
    print(f"Approved flips collected: {len(approved)}")
    if not approved:
        print("Nothing to apply.")
        return

    by_sheet = {}
    for a in approved:
        by_sheet.setdefault(a["_source_sheet"], 0)
        by_sheet[a["_source_sheet"]] += 1
    for sh, n in by_sheet.items():
        print(f"  {sh}: {n}")

    if args.dry_run:
        print("\nFirst 20:")
        for a in approved[:20]:
            print(f"  row={a.get('row')} {a.get('paper_id')}: "
                  f"{a.get('current_decision')} → {a.get('recommended_decision')} "
                  f"[{a.get('source_agents')}]")
        if len(approved) > 20:
            print(f"  ... +{len(approved) - 20} more")
        print("\nDry-run complete. Use --apply to write.")
        return

    log_entries, skipped = apply_flips(approved)
    print(f"\nApplied {len(log_entries)} flips to {XLSX}")
    print(f"Backup: {BACKUP}")
    if skipped:
        print(f"Skipped {len(skipped)}: {skipped[:10]}{'...' if len(skipped) > 10 else ''}")
    append_log(log_entries)
    print(f"Logged to {LOG_FILE}")


if __name__ == "__main__":
    main()
