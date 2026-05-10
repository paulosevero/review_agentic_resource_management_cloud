"""Generate the '04 — Title Screening' tab in spreadsheet.xlsx.

Reads screening/title/consolidated.json and writes a tab sorted by
score_diff descending (highest disagreement first).
Columns:
  Paper ID | Title | Sub-Agent 1 Score | Sub-Agent 1 Decision |
  Sub-Agent 2 Score | Sub-Agent 2 Decision | Final Score |
  Disagreement Type | Machine Decision | My Final Decision | My Justification
"""

import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = "/Users/paulosouza/Documents/Carreira/Pesquisa/Código/review_agentic_resource_management_cloud"
CONSOLIDATED_PATH = os.path.join(REPO_ROOT, "screening", "title", "consolidated.json")
XLSX_PATH = os.path.join(REPO_ROOT, "spreadsheet.xlsx")
SHEET_NAME = "04 — Title Screening"

HEADERS = [
    "Paper ID",
    "Title",
    "Sub-Agent 1 Score",
    "Sub-Agent 1 Decision",
    "Sub-Agent 2 Score",
    "Sub-Agent 2 Decision",
    "Final Score",
    "Difference in Final Score Between Sub-Agents",
    "Disagreement Type",
    "Machine Decision",
    "My Final Decision",
    "My Justification",
]

COL_WIDTHS = [12, 80, 18, 20, 18, 20, 12, 44, 30, 18, 20, 40]


def main():
    """Generate the title screening tab."""
    with open(CONSOLIDATED_PATH, encoding="utf-8") as f:
        data = json.load(f)

    # Sort by score_diff descending (highest disagreement first)
    data_sorted = sorted(data, key=lambda x: -x["score_diff"])

    wb = openpyxl.load_workbook(XLSX_PATH)

    # Remove existing sheet if it exists
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(title=SHEET_NAME)

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    include_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    exclude_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    override_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    disagree_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for row_idx, d in enumerate(data_sorted, start=2):
        row_data = [
            d["paper_id"],
            d["title"],
            d["sub_agent_1_score"],
            d["sub_agent_1_decision"],
            d["sub_agent_2_score"],
            d["sub_agent_2_decision"],
            d["final_score"],
            d["score_diff"],
            d["disagreement_type"],
            d["machine_decision"],
            "",  # My Final Decision
            "",  # My Justification
        ]

        # Choose row fill based on disagreement type
        if d["disagreement_type"] == "strong_disagreement":
            row_fill = disagree_fill
        elif d["disagreement_type"] == "doctrine_override_c3_absent":
            row_fill = override_fill
        elif d["machine_decision"] == "include":
            row_fill = include_fill
        else:
            row_fill = exclude_fill

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            if col_idx == 2:  # Title column
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(XLSX_PATH)
    print(f"Saved '{SHEET_NAME}' tab to {XLSX_PATH}")
    print(f"  Rows: {len(data_sorted)}")
    include_count = sum(1 for d in data if d["machine_decision"] == "include")
    exclude_count = sum(1 for d in data if d["machine_decision"] == "exclude")
    print(f"  Machine include: {include_count}, Machine exclude: {exclude_count}")


if __name__ == "__main__":
    main()
