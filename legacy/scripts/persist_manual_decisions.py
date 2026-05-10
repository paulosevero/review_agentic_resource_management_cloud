"""Imports My Final Decision / My Justification from the spreadsheet into paper frontmatter.

Reads `spreadsheet_title_screening.xlsx`, aba `04 — Title Screening`. For each row,
locates the corresponding `papers/paper-NNN.md` file and updates:
  - status["04-title-screening"] → lowercased decision (include/exclude/pending)
  - screening["04-title-screening"]["human_decision"] → lowercased decision
  - screening["04-title-screening"]["human_justification"] → the justification string

Uses targeted regex replacement to preserve YAML frontmatter formatting (quotes,
indentation, key order). Idempotent — re-runs are safe.
"""

import argparse
import re
from pathlib import Path

import openpyxl


def update_paper(content: str, decision_lower: str, justification: str) -> str:
    """Updates the paper file content with the user's manual decision and justification.

    Args:
        content (str): Original file content.
        decision_lower (str): Lowercased decision ("include", "exclude", or "pending").
        justification (str): Justification string from My Justification column.

    Returns:
        new_content (str): Updated file content.
    """
    escaped_just = justification.replace("\\", "\\\\").replace('"', '\\"')
    new_content = re.sub(
        pattern=r'(\n  "04-title-screening":)\s+\w+',
        repl=rf'\1 {decision_lower}',
        string=content,
        count=1,
    )
    new_content = re.sub(
        pattern=r'(    human_decision:)\s+"[^"]*"',
        repl=rf'\1 "{decision_lower}"',
        string=new_content,
        count=1,
    )
    new_content = re.sub(
        pattern=r'(    human_justification:)\s+"[^"]*"',
        repl=rf'\1 "{escaped_just}"',
        string=new_content,
        count=1,
    )
    return new_content


def load_decisions(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Loads paper_id, my_final_decision, my_justification from the spreadsheet.

    Args:
        xlsx_path (Path): Path to spreadsheet.xlsx.
        sheet_name (str): Sheet name.

    Returns:
        decisions (list[dict]): One dict per paper.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    pid_col = headers.index("Paper ID")
    decision_col = headers.index("My Final Decision")
    just_col = headers.index("My Justification")
    decisions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        decisions.append({
            "paper_id": row[pid_col],
            "decision": (row[decision_col] or "").strip(),
            "justification": (row[just_col] or "").strip(),
        })
    return decisions


def persist_to_paper_files(decisions: list[dict], papers_dir: Path) -> dict:
    """Writes each decision to the matching paper-NNN.md file.

    Args:
        decisions (list[dict]): Output of load_decisions.
        papers_dir (Path): Directory containing paper-NNN.md files.

    Returns:
        summary (dict): Counts of updates, skips, errors.
    """
    updated_count = 0
    skipped_empty_count = 0
    not_found_count = 0
    for d in decisions:
        if not d["decision"]:
            skipped_empty_count += 1
            continue
        paper_path = papers_dir / f"{d['paper_id']}.md"
        if not paper_path.exists():
            not_found_count += 1
            print(f"WARNING: {paper_path} not found")
            continue
        original = paper_path.read_text(encoding="utf-8")
        decision_lower = d["decision"].lower()
        new_content = update_paper(content=original, decision_lower=decision_lower, justification=d["justification"])
        if new_content != original:
            paper_path.write_text(new_content, encoding="utf-8")
            updated_count += 1
    summary = {
        "total_rows": len(decisions),
        "updated": updated_count,
        "skipped_empty": skipped_empty_count,
        "not_found": not_found_count,
    }
    return summary


def main(xlsx_path: str, papers_dir: str, sheet_name: str) -> None:
    decisions = load_decisions(xlsx_path=Path(xlsx_path), sheet_name=sheet_name)
    summary = persist_to_paper_files(decisions=decisions, papers_dir=Path(papers_dir))
    print(summary)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="spreadsheet_title_screening.xlsx")
    parser.add_argument("--papers-dir", default="papers")
    parser.add_argument("--sheet", default="04 — Title Screening")
    args = parser.parse_args()
    main(xlsx_path=args.xlsx, papers_dir=args.papers_dir, sheet_name=args.sheet)
