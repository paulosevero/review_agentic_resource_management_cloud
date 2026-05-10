"""Normalizing paper IDs to 4-digit zero-padded form (paper-NNNN).

Normalizes:
- All sheets of Análise.xlsx that have a 'Paper ID' column.
- Filenames in raw/pdfs/ matching pattern '[paper-N+] <title>.<ext>'.

Usage:
    python 01_normalize_ids.py --dry-run    # preview only
    python 01_normalize_ids.py --apply      # mutates files in-place

Idempotent: re-running on already-normalized state is a no-op.
"""

import argparse
import re
import shutil
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
XLSX_PATH = REPO_ROOT / "Análise.xlsx"
PDFS_DIR = REPO_ROOT / "raw" / "pdfs"

ID_PATTERN_TEXT = re.compile(r"^paper-(\d+)$")
ID_PATTERN_FILENAME = re.compile(r"^\[paper-(\d+)\](.*)$")


def normalize_id(raw_id: str) -> str:
    """Returning 'paper-NNNN' form. Returns input unchanged if it doesn't match."""
    m = ID_PATTERN_TEXT.match(str(raw_id).strip())
    if not m:
        return raw_id
    return f"paper-{int(m.group(1)):04d}"


def normalize_xlsx(dry_run: bool) -> dict:
    """Rewriting every 'Paper ID' cell across all sheets to paper-NNNN form."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    changes_per_sheet = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Accepting both 'Paper ID' (screening tabs) and 'id' (Corpus, 07-Taxonomy).
        id_col_name = next((n for n in ("Paper ID", "id") if n in header), None)
        if id_col_name is None:
            continue
        col_idx = header.index(id_col_name) + 1  # 1-based for openpyxl
        changes = 0
        for row in ws.iter_rows(min_row=2):
            cell = row[col_idx - 1]
            old = cell.value
            if old is None:
                continue
            new = normalize_id(old)
            if new != old:
                changes += 1
                if not dry_run:
                    cell.value = new
        changes_per_sheet[sheet_name] = changes
    if not dry_run:
        wb.save(XLSX_PATH)
    return changes_per_sheet


def normalize_pdfs(dry_run: bool) -> list:
    """Renaming files in raw/pdfs/ from '[paper-N] X' to '[paper-NNNN] X'."""
    if not PDFS_DIR.exists():
        return []
    renames = []
    for entry in sorted(PDFS_DIR.iterdir()):
        if not entry.is_file():
            continue
        m = ID_PATTERN_FILENAME.match(entry.name)
        if not m:
            continue
        old_id_num = m.group(1)
        rest = m.group(2)
        new_name = f"[paper-{int(old_id_num):04d}]{rest}"
        if new_name == entry.name:
            continue
        new_path = entry.with_name(new_name)
        renames.append((entry.name, new_name))
        if not dry_run:
            entry.rename(new_path)
    return renames


def main():
    ap = argparse.ArgumentParser()
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--dry-run", action="store_true")
    g.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    dry = args.dry_run

    print(f"=== Análise.xlsx — {'DRY-RUN' if dry else 'APPLY'} ===")
    xlsx_changes = normalize_xlsx(dry_run=dry)
    for sheet, n in xlsx_changes.items():
        print(f"  {sheet!r}: {n} cells normalized")

    print(f"\n=== raw/pdfs/ — {'DRY-RUN' if dry else 'APPLY'} ===")
    renames = normalize_pdfs(dry_run=dry)
    print(f"  total renames: {len(renames)}")
    for old, new in renames[:20]:
        print(f"    {old}  ->  {new}")
    if len(renames) > 20:
        print(f"    ... ({len(renames) - 20} more)")


if __name__ == "__main__":
    main()
