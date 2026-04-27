"""Writes the stage-05 classifier decisions (with tiebreaker overrides) into spreadsheet.xlsx.

Creates or replaces the `05 — Abstract Screening` tab with:
  - Paper ID, Title, Abstract, Auto Decision, Auto Justification, Winning Category,
    Evidence, My Final Decision (blank for user), My Justification (blank for user)
  - Background fill: green (#E7FFEE) on Paper ID + Title for Include,
    red (#FFEFF0) for Exclude
  - Sort: Excludes first (red block — classical_agents at top for spot-check),
    then Includes (green block — mas_llm_based / agent_llm_based first, then generic)
  - Rich-text bold for matched evidence in Title and Abstract cells

Idempotent: re-running replaces the tab in place. The user's prior `My Final Decision`
edits are preserved if present (matched by Paper ID).
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

GREEN_FILL = PatternFill(start_color="FFE7FFEE", end_color="FFE7FFEE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFFFEFF0", end_color="FFFFEFF0", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
BOLD_FONT = InlineFont(b=True)
HEADERS = [
    "Paper ID", "Title", "Abstract", "Auto Decision", "Auto Justification",
    "Winning Category", "Evidence", "My Final Decision", "My Justification",
]
EXCLUDE_ORDER = ["out_of_scope", "review", "classical_agents", "rl", "llm_as_workload", "no_signal"]
INCLUDE_ORDER = ["mas_llm_based", "agent_llm_based", "llm_agentic_ai_generic"]


def load_full_abstract(paper_id: str, papers_dir: Path) -> str:
    """Reads the full abstract from a paper file (the JSON has only a preview)."""
    path = papers_dir / f"{paper_id}.md"
    if not path.exists():
        return ""
    content = path.read_text(encoding="utf-8")
    m = re.search(pattern=r'###\s*Abstract\s*\n+(.+?)(?=\n###|\n##|\Z)', string=content, flags=re.DOTALL)
    return m.group(1).strip() if m else ""


def load_existing_my_decisions(wb: openpyxl.Workbook, sheet_name: str) -> dict:
    """If the sheet already exists, reads the user's My Final Decision and Justification per paper_id."""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    if "Paper ID" not in headers:
        return {}
    pid_col = headers.index("Paper ID")
    dec_col = headers.index("My Final Decision") if "My Final Decision" in headers else None
    just_col = headers.index("My Justification") if "My Justification" in headers else None
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[pid_col]
        if not pid:
            continue
        out[pid] = {
            "decision": row[dec_col] if dec_col is not None else None,
            "justification": row[just_col] if just_col is not None else None,
        }
    return out


def make_rich_text_with_bold(text: str, bold_substrings: list[str]) -> CellRichText | str:
    """Builds a CellRichText with the first occurrence of each substring bolded.

    Args:
        text (str): Source text.
        bold_substrings (list[str]): Distinct substrings to bold.

    Returns:
        result: CellRichText if any bolding applied, else the plain string.
    """
    if not text or not bold_substrings:
        return text
    spans = []
    for sub in bold_substrings:
        if not sub:
            continue
        m = re.search(pattern=re.escape(sub), string=text, flags=re.IGNORECASE)
        if m:
            spans.append((m.start(), m.end()))
    if not spans:
        return text
    spans.sort(key=lambda s: s[0])
    merged = [spans[0]]
    for start, end in spans[1:]:
        if start <= merged[-1][1]:
            merged[-1] = (merged[-1][0], max(end, merged[-1][1]))
        else:
            merged.append((start, end))
    blocks = []
    cursor = 0
    for start, end in merged:
        if start > cursor:
            blocks.append(text[cursor:start])
        blocks.append(TextBlock(font=BOLD_FONT, text=text[start:end]))
        cursor = end
    if cursor < len(text):
        blocks.append(text[cursor:])
    return CellRichText(*blocks)


def category_sort_key(decision: dict) -> tuple:
    """Sort key: Excludes first, then Includes; within each block, by category order then paper_id."""
    cat = decision["winning_category"]
    if decision["decision"] == "Exclude":
        return (0, EXCLUDE_ORDER.index(cat) if cat in EXCLUDE_ORDER else 999, decision["paper_id"])
    return (1, INCLUDE_ORDER.index(cat) if cat in INCLUDE_ORDER else 999, decision["paper_id"])


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, decisions: list[dict],
                papers_dir: Path, my_decisions: dict) -> None:
    """Creates/replaces the screening sheet and populates it with formatted rows."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    decisions_sorted = sorted(decisions, key=category_sort_key)
    for row_idx, d in enumerate(decisions_sorted, start=2):
        full_abstract = load_full_abstract(paper_id=d["paper_id"], papers_dir=papers_dir)
        evidence_substrings = list({e["matched_text"] for e in d["evidence"]})
        title_cell_value = make_rich_text_with_bold(text=d["title"], bold_substrings=evidence_substrings)
        abstract_cell_value = make_rich_text_with_bold(text=full_abstract, bold_substrings=evidence_substrings)
        evidence_summary = ", ".join(f"{e['category']}/{e['pattern_id']}={e['matched_text'][:30]}" for e in d["evidence"][:5])
        ws.cell(row=row_idx, column=1, value=d["paper_id"])
        ws.cell(row=row_idx, column=2).value = title_cell_value
        ws.cell(row=row_idx, column=3).value = abstract_cell_value
        ws.cell(row=row_idx, column=4, value=d["decision"])
        ws.cell(row=row_idx, column=5, value=d["justification"])
        ws.cell(row=row_idx, column=6, value=d["winning_category"])
        ws.cell(row=row_idx, column=7, value=evidence_summary)
        prior = my_decisions.get(d["paper_id"], {})
        ws.cell(row=row_idx, column=8, value=prior.get("decision"))
        ws.cell(row=row_idx, column=9, value=prior.get("justification"))
        fill = GREEN_FILL if d["decision"] == "Include" else RED_FILL
        ws.cell(row=row_idx, column=1).fill = fill
        ws.cell(row=row_idx, column=2).fill = fill
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    column_widths = {1: 12, 2: 50, 3: 90, 4: 12, 5: 35, 6: 22, 7: 60, 8: 16, 9: 28}
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def apply_tiebreaker_overrides(decisions: list[dict], tiebreaker_path: Path) -> list[dict]:
    """Replaces decisions for papers in the tiebreaker results file."""
    if not tiebreaker_path.exists():
        return decisions
    with open(tiebreaker_path, "r", encoding="utf-8") as f:
        overrides = json.load(f)
    by_id = {o["paper_id"]: o for o in overrides}
    for d in decisions:
        if d["paper_id"] in by_id:
            o = by_id[d["paper_id"]]
            d["decision"] = o["decision"]
            d["justification"] = f"[Tiebreaker LLM, conf={o.get('confidence', '?')}] {o['justification']}"
    return decisions


def main(xlsx_path: str, decisions_path: str, tiebreaker_path: str, papers_dir: str, sheet_name: str) -> None:
    with open(decisions_path, "r", encoding="utf-8") as f:
        decisions = json.load(f)
    decisions = apply_tiebreaker_overrides(decisions=decisions, tiebreaker_path=Path(tiebreaker_path))
    wb = openpyxl.load_workbook(xlsx_path)
    my_decisions = load_existing_my_decisions(wb=wb, sheet_name=sheet_name)
    write_sheet(wb=wb, sheet_name=sheet_name, decisions=decisions,
                papers_dir=Path(papers_dir), my_decisions=my_decisions)
    wb.save(xlsx_path)
    counts = {"Include": 0, "Exclude": 0}
    for d in decisions:
        counts[d["decision"]] = counts.get(d["decision"], 0) + 1
    print(f"Wrote {len(decisions)} rows into '{sheet_name}': {counts}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="spreadsheet.xlsx")
    parser.add_argument("--decisions", default="screening/abstract/classifier-decisions.json")
    parser.add_argument("--tiebreaker", default="screening/abstract/tiebreaker-decisions.json")
    parser.add_argument("--papers-dir", default="papers")
    parser.add_argument("--sheet", default="05 — Abstract Screening")
    args = parser.parse_args()
    main(xlsx_path=args.xlsx, decisions_path=args.decisions, tiebreaker_path=args.tiebreaker,
         papers_dir=args.papers_dir, sheet_name=args.sheet)
