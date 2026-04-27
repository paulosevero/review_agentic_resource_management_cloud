"""Audits the deterministic title-stage classifier against the user's manual decisions.

Reads `My Final Decision` from `spreadsheet.xlsx` (aba `04 — Title Screening`),
runs the classifier from `protocols/classifier-config-title.json` on each title,
and reports stratified accuracy and a confusion matrix. Sample disagreements
are dumped into a Markdown report so the user can spot under- and over-fitting
of the regex rules. The classifier output never overwrites the user's manual
decisions — this is read-only audit.
"""

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from deterministic_classifier import classify, load_config


def load_spreadsheet_decisions(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Loads paper_id, title, and the manual decision from the screening spreadsheet.

    Args:
        xlsx_path (Path): Path to spreadsheet.xlsx.
        sheet_name (str): Name of the screening tab.

    Returns:
        rows (list[dict]): One dict per paper with paper_id, title, manual_decision.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    pid_col = headers.index("Paper ID")
    title_col = headers.index("Title")
    decision_col = headers.index("My Final Decision")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({
            "paper_id": row[pid_col],
            "title": row[title_col] or "",
            "manual_decision": (row[decision_col] or "").strip(),
        })
    return rows


def run_audit(xlsx_path: Path, config_path: Path, out_path: Path) -> dict:
    """Runs the audit, writes the Markdown report, and returns a summary dict.

    Args:
        xlsx_path (Path): Path to spreadsheet.xlsx.
        config_path (Path): Path to the title-stage classifier config JSON.
        out_path (Path): Path to write the Markdown report.

    Returns:
        summary (dict): Counts, match rate, and confusion matrix.
    """
    config = load_config(path=str(config_path))
    rows = load_spreadsheet_decisions(xlsx_path=xlsx_path, sheet_name="04 — Title Screening")
    confusion = Counter()
    by_winning_category = defaultdict(lambda: Counter())
    disagreements = []
    for r in rows:
        result = classify(text=r["title"], config=config)
        manual = r["manual_decision"]
        classifier = result["decision"]
        confusion[(manual, classifier)] += 1
        by_winning_category[result["winning_category"]][(manual, classifier)] += 1
        if manual and manual != classifier:
            disagreements.append({
                "paper_id": r["paper_id"],
                "title": r["title"],
                "manual": manual,
                "classifier": classifier,
                "winning_category": result["winning_category"],
                "justification": result["justification"],
                "evidence": [m["matched_text"] for m in result["evidence"][:5]],
            })
    total = sum(confusion.values())
    matches = confusion[("Include", "Include")] + confusion[("Exclude", "Exclude")]
    match_rate = matches / total if total else 0.0
    user_includes = confusion[("Include", "Include")] + confusion[("Include", "Exclude")]
    user_excludes = confusion[("Exclude", "Include")] + confusion[("Exclude", "Exclude")]
    include_recall = confusion[("Include", "Include")] / user_includes if user_includes else 0.0
    exclude_recall = confusion[("Exclude", "Exclude")] / user_excludes if user_excludes else 0.0
    write_report(
        out_path=out_path,
        total=total,
        matches=matches,
        match_rate=match_rate,
        confusion=confusion,
        by_winning_category=by_winning_category,
        disagreements=disagreements,
        include_recall=include_recall,
        exclude_recall=exclude_recall,
    )
    summary = {
        "total": total,
        "matches": matches,
        "match_rate": match_rate,
        "include_recall": include_recall,
        "exclude_recall": exclude_recall,
        "confusion": {f"manual={k[0]}|cls={k[1]}": v for k, v in confusion.items()},
        "false_negatives": confusion[("Include", "Exclude")],
        "false_positives": confusion[("Exclude", "Include")],
        "report_path": str(out_path),
    }
    return summary


def write_report(out_path: Path, total: int, matches: int, match_rate: float, confusion: Counter,
                 by_winning_category: dict, disagreements: list, include_recall: float, exclude_recall: float) -> None:
    """Writes the stratified audit report to disk in Markdown format.

    Args:
        out_path (Path): Where to save.
        total (int): Total papers audited.
        matches (int): Papers where classifier and manual agree.
        match_rate (float): matches/total.
        confusion (Counter): Counter of (manual, classifier) tuples.
        by_winning_category (dict): Per-winning-category breakdown of (manual, classifier).
        disagreements (list): List of disagreement records.
        include_recall (float): Of user's Includes, fraction the classifier also included.
        exclude_recall (float): Of user's Excludes, fraction the classifier also excluded.

    Returns:
        None.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines = []
    lines.append("# Stage 04 — Classifier vs. Manual Audit\n")
    lines.append(f"- Total papers: **{total}**")
    lines.append(f"- Overall match: **{matches}/{total} ({match_rate:.1%})**")
    lines.append(f"- Include recall (classifier finds user's Includes): **{include_recall:.1%}**")
    lines.append(f"- Exclude recall (classifier matches user's Excludes): **{exclude_recall:.1%}**")
    lines.append("")
    lines.append("## Confusion matrix\n")
    lines.append("| Manual \\ Classifier | Include | Exclude |")
    lines.append("|---|---|---|")
    lines.append(f"| **Include** | {confusion[('Include', 'Include')]} | {confusion[('Include', 'Exclude')]} (false negatives) |")
    lines.append(f"| **Exclude** | {confusion[('Exclude', 'Include')]} (false positives) | {confusion[('Exclude', 'Exclude')]} |")
    lines.append("")
    lines.append("## By winning category\n")
    lines.append("| Category | Total | Manual=Include | Manual=Exclude | Match% |")
    lines.append("|---|---:|---:|---:|---:|")
    for cat in sorted(by_winning_category.keys()):
        c = by_winning_category[cat]
        ttl = sum(c.values())
        m_inc = c[("Include", "Include")] + c[("Include", "Exclude")]
        m_exc = c[("Exclude", "Include")] + c[("Exclude", "Exclude")]
        agree = c[("Include", "Include")] + c[("Exclude", "Exclude")]
        rate = agree / ttl if ttl else 0
        lines.append(f"| {cat} | {ttl} | {m_inc} | {m_exc} | {rate:.1%} |")
    lines.append("")
    fn = [d for d in disagreements if d["manual"] == "Include" and d["classifier"] == "Exclude"]
    fp = [d for d in disagreements if d["manual"] == "Exclude" and d["classifier"] == "Include"]
    lines.append(f"## False negatives ({len(fn)}) — user said Include, classifier said Exclude\n")
    lines.append("These are the papers the classifier MISSED. Tightening rules should focus on adding patterns that catch these.\n")
    lines.append("| Paper ID | Title | Winning Category | Evidence |")
    lines.append("|---|---|---|---|")
    for d in fn[:50]:
        ev = ", ".join(d["evidence"]) if d["evidence"] else "(none)"
        title_clean = (d["title"] or "").replace("|", "\\|")[:120]
        lines.append(f"| {d['paper_id']} | {title_clean} | {d['winning_category']} | {ev} |")
    if len(fn) > 50:
        lines.append(f"\n_…and {len(fn) - 50} more false negatives._\n")
    lines.append("")
    lines.append(f"## False positives ({len(fp)}) — user said Exclude, classifier said Include\n")
    lines.append("These are the papers the classifier OVER-INCLUDED. Adding overrides or out-of-scope patterns can fix.\n")
    lines.append("| Paper ID | Title | Winning Category | Evidence |")
    lines.append("|---|---|---|---|")
    for d in fp[:50]:
        ev = ", ".join(d["evidence"]) if d["evidence"] else "(none)"
        title_clean = (d["title"] or "").replace("|", "\\|")[:120]
        lines.append(f"| {d['paper_id']} | {title_clean} | {d['winning_category']} | {ev} |")
    if len(fp) > 50:
        lines.append(f"\n_…and {len(fp) - 50} more false positives._\n")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def main(xlsx_path: str, config_path: str, out_path: str) -> None:
    summary = run_audit(
        xlsx_path=Path(xlsx_path),
        config_path=Path(config_path),
        out_path=Path(out_path),
    )
    print(json.dumps(summary, indent=2, default=str))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="spreadsheet.xlsx")
    parser.add_argument("--config", default="protocols/classifier-config-title.json")
    parser.add_argument("--out", default="audits/stage-04-classifier-vs-manual.md")
    args = parser.parse_args()
    main(xlsx_path=args.xlsx, config_path=args.config, out_path=args.out)
